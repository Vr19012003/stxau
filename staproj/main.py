import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import yaml

# Load user credentials from config.yaml
def load_credentials():
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
    return config['users']

# Save new user credentials to the config.yaml file
def save_credentials(new_user):
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
    
    config['users'].update(new_user)

    with open('config.yaml', 'w') as file:
        yaml.dump(config, file)

# Check if the user is authenticated and has the appropriate role
def authenticate(username, password):
    users = load_credentials()
    if username in users and users[username]['password'] == password:
        return users[username]['role']
    return None

# Define a function to store data in an Excel file
def save_to_excel(data, filename='form_data.xlsx'):
    df = pd.DataFrame([data])

    if not os.path.exists(filename):
        # Create a new Excel file if it doesn't exist
        df.to_excel(filename, index=False)
    else:
        try:
            # Load the existing workbook
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book

            # Determine the last row in the existing sheet
            startrow = book['Sheet1'].max_row

            # Append the new data
            df.to_excel(writer, index=False, header=False, startrow=startrow)
            writer.save()
        except Exception as e:
            st.error(f"An error occurred: {e}")

# Function to load and display Excel data
def load_and_display_excel(filename='form_data.xlsx'):
    if os.path.exists(filename):
        df = pd.read_excel(filename)
        st.dataframe(df)  # Display the DataFrame in Streamlit
    else:
        st.warning("No data found. Please submit the form first.")

# Streamlit App UI
st.title("User Input Form with Role-Based Access")

# Authentication state management
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False
    st.session_state['role'] = None

# Choose role: Admin or User
role_choice = st.sidebar.selectbox("Select your role", ["Admin", "User"])

# Admin Login
if role_choice == "Admin":
    if not st.session_state['authenticated']:
        st.subheader("Admin Login")
        username = st.text_input("Admin Username")
        password = st.text_input("Admin Password", type="password")
        admin_login_button = st.button("Login as Admin")

        if admin_login_button:
            role = authenticate(username, password)
            if role == "admin":
                st.session_state['authenticated'] = True
                st.session_state['role'] = role
                st.success("Admin logged in successfully!")
            else:
                st.error("Invalid Admin credentials.")

# User Sign-up/Login
elif role_choice == "User":
    if not st.session_state['authenticated']:
        # Provide Sign-up and Login options for Users
        action = st.radio("Do you want to Sign-up or Login?", ["Login", "Sign-up"])

        if action == "Login":
            st.subheader("User Login")
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            user_login_button = st.button("Login")

            if user_login_button:
                role = authenticate(username, password)
                if role == "user":
                    st.session_state['authenticated'] = True
                    st.session_state['role'] = role
                    st.success("Logged in successfully as a user!")
                else:
                    st.error("Invalid user credentials.")

        elif action == "Sign-up":
            st.subheader("User Sign-up")
            new_username = st.text_input("New Username")
            new_password = st.text_input("New Password", type="password")
            user_signup_button = st.button("Sign Up")

            if user_signup_button:
                if new_username and new_password:
                    users = load_credentials()
                    if new_username in users:
                        st.error("Username already exists!")
                    else:
                        # Store new user credentials with default role 'user'
                        new_user = {
                            new_username: {
                                'username': new_username,
                                'password': new_password,
                                'role': 'user'
                            }
                        }
                        save_credentials(new_user)
                        st.success("Sign-up successful! You can now log in.")
                else:
                    st.error("Please fill in both username and password.")

# If authenticated, show the form based on the role
if st.session_state['authenticated']:
    st.subheader(f"Logged in as {st.session_state['role']}")

    # Admin Access: Can read and write data
    if st.session_state['role'] == 'admin':
        st.write("**Admin Access: You can both submit and view data.**")

        # Create a form for Admin to enter data
        with st.form(key="admin_form"):
            name = st.text_input("Enter your name")
            gender = st.radio("Select your gender", ("Male", "Female", "Other"))
            submit_button = st.form_submit_button(label="Submit")

        # Handle form submission
        if submit_button:
            if name and gender:
                data = {"Name": name, "Gender": gender}
                save_to_excel(data)
                st.success(f"Data saved! Name: {name}, Gender: {gender}")
            else:
                st.error("Please fill out the form completely.")
        
        # Display data
        if st.button("Show Data"):
            load_and_display_excel()

    # User Access: Can only read data
    elif st.session_state['role'] == 'user':
        st.write("**User Access: You can only view data.**")
        load_and_display_excel()
